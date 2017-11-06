

"""

Helpful Stack Overflow:
https://stackoverflow.com/questions/22813814/clearly-documented-reading-of-emails-functionality-with-python-win32com-outlook/35801030#35801030
https://stackoverflow.com/questions/22399835/how-to-save-attachment-from-outlook-using-win32com-client-in-python

"""

import win32com.client
import re
import os
import openpyxl
import time
from datetime import datetime, timedelta


start = time.time()

print('Starting the script...\n')

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

destination = "C:\\Users\\X834995\\PycharmProjects\\OutlookScraping\\Craig Nangle Files"
complete_destination = "C:\\Users\\X834995\\PycharmProjects\\OutlookScraping\\Craig Nangle Files\\Completed Files"
directory = os.fsencode(destination)
complete_directory = os.fsencode(complete_destination)

N = 22  # Number of days to look back for email
date_N_days_ago = datetime.now() - timedelta(days=N)


print("Searching outlook email for Craig's email with attachment...")
print("After:", str(date_N_days_ago), "\n")


inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.Sort("ReceivedTime", True)  # win32com Sort method defaults to Ascending (False). Change to True to be desc.

counter = 0

# --------------------------------------------------------------------------------------------
# Go through all of the messages, find if Craig sent them, and then download the attachment
# --------------------------------------------------------------------------------------------

for m in messages:
    sender = m.SenderEmailAddress.upper()
    received_date = str(m.ReceivedTime)[0:16]  # get the date in which we received the email found above
    received_date = datetime.strptime(received_date, '%Y-%m-%d %H:%M')
    if received_date >= date_N_days_ago:  # Only doing below for the items N days ago
        if re.search('NANGLE', sender) is not None:  # Need to upper and regex because internal emails look strange
            print(received_date)
            print(sender, m.ReceivedTime, "\n")
            for a in m.Attachments:
                if re.search('.xlsx', a.FileName.lower()):
                    counter += 1
                    print(a.FileName, "\n")
                    a.SaveAsFile(destination + "\\" + a.FileName)
    else:
        break

print("Found", str(counter), "file(s)")

# -----------------------------------------------------------------------------------------------
# Iterating through the files in main directory
# https://stackoverflow.com/questions/10377998/how-can-i-iterate-over-files-in-a-given-directory
# -----------------------------------------------------------------------------------------------

print("\nLooping through files and creating queries...\n")
filenames = []

# Find only the .xlsx files
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith(".xlsx"):
        filename = os.path.join(destination, filename)
        print(os.path.join(destination, filename))
        filenames.append(filename)
        continue
    else:
        continue

# new_file = re.sub('.xlsx', '\\\\', filename)  # Potentially rename directory

# Go through the xlsx files, open them, and then
for i in range(len(filenames)):

    wb = openpyxl.load_workbook(filenames[i], read_only=True)
    ws = wb.active

    # First Row = Column Names
    colrange = ws[1]  # First row of the worksheet == columm headers
    print(filenames[i])

    # Loop through the column headers to find "VIN", if not, it's just the first column
    for j in range(len(colrange)):
        if "VIN" in colrange[j].value.upper():
            # print(j)
            # print(colrange[j].value)
            vin_column = i + 1  # Have to add 1 because index starts at 0, but excel starts at 1
            # break

            # Loop through the rows in the column (minus row 1 b/c it's a header), and create the subsql statement
            for k, row in enumerate(ws.iter_rows(min_row=2, min_col=vin_column, max_col=vin_column)):
                for cell in row:
                    if k == 0:
                        # print("SELECT '", str(cell.value), "' AS VIN_ID FROM DUAL", sep="")
                        subsql = "SELECT '" + str(cell.value) + "' AS VIN_ID FROM DUAL"
                    else:
                        # print("UNION ALL SELECT '", str(cell.value), "' FROM DUAL", sep="")
                        subsql += "\nUNION ALL SELECT '" + str(cell.value) + "' FROM DUAL"

        else:
            vin_column = 1  # Assuming first column is the VIN

            for k, row in enumerate(ws.iter_rows(min_row=1, min_col=vin_column, max_col=vin_column)):
                for cell in row:
                    if k == 0:
                        # print("SELECT '", str(cell.value), "' AS VIN_ID FROM DUAL", sep="")
                        subsql = "SELECT '" + str(cell.value) + "' AS VIN_ID FROM DUAL"

                    if str(cell.value) != 'None':  # If the cell value is not blank, add the VIN
                        # print("UNION ALL SELECT '", str(cell.value), "' FROM DUAL", sep="")
                        subsql += "\nUNION ALL SELECT '" + str(cell.value) + "' FROM DUAL"

                    else:
                        break  # If the cell value is blank, exit the loop (much faster)
                        # print("UNION ALL SELECT '", str(cell.value), "' FROM DUAL", sep="")
                        # subsql += "\nUNION ALL SELECT '" + str(cell.value) + "' FROM DUAL"


    # Once we have the subsql, now we can write the full SQL statement to write to the .txt file

    sql = """ SELECT
        
          DLR.NMAC_RGN_CD AS "REGION_CODE"
          ,SD.SLS_ARA_CD AS "AREA_CODE"
          ,SA.ARA_NM AS "AREA_NAME"
          ,SD.DSTRCT_CD AS "DISTRICT_CODE"
          ,DLR.DLR_NB AS "DEALER_CODE"
          ,DLR.DLR_NM AS "DEALER_NAME"
          ,DI.VIN_ID
          ,LS.VHCL_LCTN_STS_CD AS "LOCATION_STATUS_CODE"
          ,LS.VHCL_LCTN_STS_NM AS "LOCATION_STATUS_DESCRIPTION"
          ,DI.RMVD_IN AS "REMOVED_INDICATOR"
          --,VC.CMPGN_NB AS "CAMPAIGN_NUMBER"
          --,CM.CMPGN_DS AS "CAMPAIGN_DESC"
          --,CM.EXTRNL_EFCTV_DT AS "EXT_EFFECTIVE_DATE"
          --,CM.CMPGN_TYP_NM AS "CAMPAIGN_TYPE"
          --,'NEW' AS "INVENTORY_SOURCE"
         -- ,CASE WHEN VC.CMPGN_NB IS NULL THEN 'No Campaign'
         --       ELSE 'Campaign' END AS "CAMPAIGN_FLAG"
          --,DI.*
               
         -- COUNT(*)
          FROM DLR_VHCL_IVNTRY DI
            
            INNER JOIN VHCL_LCTN_STS LS
              ON DI.LCTN_STS_CD = LS.VHCL_LCTN_STS_CD
            /*
            LEFT JOIN VHCL_CMPGN VC -- GETTING VIN LEVEL CAMPAIGN STATUS
              ON DI.VIN_ID = VC.VIN_ID
              AND VC.CMPGN_STS_CD = 'O' -- OPEN CAMPAIGNS ONLY
              --AND VC.NNA_STS_CD = '0' -- CAMPAIGN ITSELF BEING CLOSED
            
            LEFT JOIN CMPGN_MSTR CM -- GETTING DETAILS ABOUT THE CAMPAIGN
              ON CM.CMPGN_NB = VC.CMPGN_NB
             */ 
            INNER JOIN DLR DLR
              ON DI.RFRNC_DLR_NB = DLR.RFRNC_DLR_NB
             -- AND DLR.NMAC_RGN_CD IS NOT NULL -- TAKING OUT NULL REGION CODES
             -- AND DLR.DLR_STS_IN = 'Y'
            
            INNER JOIN SLS_DSTRCT SD
              ON SD.SLS_DSTRCT_CD = DLR.SLS_DSTRCT_CD
          
            INNER JOIN SLS_ARA SA
              ON SD.SLS_ARA_CD = SA.SLS_ARA_CD
              
            INNER JOIN (
            """ + subsql + """ ) A
            ON DI.VIN_ID = A.VIN_ID"""

    text_file_path = re.sub('.xlsx', '.txt', filenames[i])  # Substituting .txt for .xlsx to get unique text files
    with open(text_file_path, "w") as text_file:  # Opening file with same name, but .txt in write mode
        text_file.write(sql)  # Writing the final SQL query to the text file.

    wb.close()

# Moving all of the files into the completed files Folder

complete_filenames = []

# Add everything in the directory to the complete list
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    #if filename.os.path.isdir():
    #    continue
    #if filename == 'Completed Files':  # Skip the "Completed Files" folder
    #    continue
    full_path = os.path.join(destination, filename)
    if os.path.isdir(full_path):
        continue
    print(os.path.join(destination, filename))
    complete_filenames.append((filename, full_path))

# Loop through files, create new path, and if the new path alraedy exists (meaning we already moved this file),
# then delete the newly downloaded version (might need to change this if he uses same document)
for file in complete_filenames:
    old_path = file[1]
    new_path = old_path.replace('C:\\Users\\X834995\\PycharmProjects\\OutlookScraping\\Craig Nangle Files\\',
                                'C:\\Users\\X834995\\PycharmProjects\\OutlookScraping\\Craig Nangle Files\\Completed Files\\')
    if os.path.isfile(new_path):
        os.remove(old_path)
    else:
        os.rename(old_path, new_path)

end = time.time()
print("\n Time to completion (seconds): ", str(end - start))
print("\nThe Script is complete!")
